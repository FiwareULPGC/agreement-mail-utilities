#!/usr/bin/env python
# -*- coding: utf-8 -*-

import getopt
import sys
import uuid

from openpyxl import *


def generate_token():
    return uuid.uuid4().hex


def generate_confirmation_tokens(filepath):

    wb = load_workbook(filepath)
    ws = wb.worksheets[0]

    ws.cell(row=1, column=4).value = "Token"
    ws.cell(row=1, column=5).value = "Sent"

    row_number = 2
    for row in ws.iter_rows(row_offset=1):
        if (row[0].value is not None or row[1].value is not None or
           row[2].value is not None):

            try:
                # Abort and notify if an empty field is found
                if row[0].value is None:
                    raise ValueError("Name value not found in row {}."
                                     .format(row_number))
                if row[1].value is None:
                    raise ValueError("Surname value not found in row {}."
                                     .format(row_number))
                if row[2].value is None:
                    raise ValueError("Email value not found in row {}."
                                     .format(row_number))

                ws.cell(row=row_number, column=4).value = generate_token()

            except Exception as e:
                ws.cell(row=row_number, column=4).value = None
                print "    Error generating token : {}".format(str(e))

            finally:
                ws.cell(row=row_number, column=5).value = 'No'

        row_number += 1

    wb.save(filepath)


if __name__ == "__main__":
    options, remainder = getopt.getopt(sys.argv[1:], 'f:h', ['file=',
                                                             'help',
                                                             ])
    example = ("Example of use: python ./generate_confirmation_tokens.py "
               "-f ./<filename>.xlsx")
    file = None
    for opt, arg in options:
        if opt in ('-f', '--file'):
            file = arg

    if file is None:
        print example
    else:
        generate_confirmation_tokens(file)
