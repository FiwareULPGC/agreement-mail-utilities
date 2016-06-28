#!/usr/bin/env python
# -*- coding: utf-8 -*-

from email.mime.text import MIMEText as text
import getopt
import smtplib
import sys

from openpyxl import *
import yaml


def send_email(smtp_host, smtp_user, smtp_pass, subject, to_address, msg):
    print "Sending e-mail to {}".format(to_address)

    # Setup SMTP server configuration
    server = smtplib.SMTP(smtp_host)
    server.starttls()
    server.login(smtp_user, smtp_pass)

    m = text(msg, 'html')
    m['Subject'] = subject
    m['From'] = smtp_user
    m['To'] = to_address

    server.sendmail(smtp_user, to_address, m.as_string())
    server.quit()


def complete_msg(template, name, surname, token):
    return template.replace('<name>', name)\
                   .replace('<surname>', surname)\
                   .replace('<token>', token)


def read_configuration(configurationpath):
    with open(configurationpath, 'r') as stream:
        return yaml.load(stream)


def send_confirmation_email(configurationpath, excelpath, resend=False):

    conf = read_configuration(configurationpath)

    wb = load_workbook(excelpath)
    ws = wb.worksheets[0]

    ws.cell(row=1, column=5).value = "Sent"

    row_number = 2
    for row in ws.iter_rows(row_offset=1):
        if (row[0].value is not None or row[1].value is not None or
           row[2].value is not None):

            try:
                to_address = str(row[2].value)

                # Abort and notify if an empty field is found
                if row[0].value is None:
                    raise ValueError("Name value not found in row {}."
                                     .format(row_number))
                if row[1].value is None:
                    raise ValueError("Surname value not found in row {}."
                                     .format(row_number))
                if row[2].value is None:
                    raise ValueError("Email value not found in row {}."
                                     .format(row_number))
                if row[3].value is None:
                    raise ValueError("Token value not found in row {}."
                                     .format(row_number))

                # If the script is in resend mode, every row will send the
                # email. If the script is not in resend mode, only the rows
                # where the sent column is not 'Yes' will send the email.
                sent_column = ws.cell(row=row_number, column=5).value
                if resend or sent_column != 'Yes':

                    msg = complete_msg(conf["template"], str(row[0].value),
                                       str(row[1].value), str(row[3].value))

                    send_email(conf["smtp_host"], conf["smtp_user"],
                               conf["smtp_pass"], conf["subject"], to_address,
                               msg)

                    ws.cell(row=row_number, column=5).value = 'Yes'

            except Exception as e:
                print "    Error sending email to {}: {}".format(to_address,
                                                                 str(e))
                ws.cell(row=row_number, column=5).value = 'No'

        row_number += 1

    wb.save(excelpath)


if __name__ == "__main__":
    options, remainder = getopt.getopt(sys.argv[1:], 'c:f:r:h', ['config',
                                                                 'file=',
                                                                 'resend',
                                                                 'help',
                                                                 ])
    example = ("Example of use:\n"
               " + Default behaviour: python ./send_confirmation_email.py "
               "-c <configfilename>.yml -f ./<filename>.xlsx\n"
               " + Force resending email: python ./send_confirmation_email.py "
               "-c <configfilename>.yml -f ./<filename>.xlsx --resend")
    config_file = None
    xlsx_file = None
    resend = False
    for opt, arg in options:
        if opt in ('-f', '--file'):
            xlsx_file = arg
        elif opt in ('-c', '--config'):
            config_file = arg
        elif opt in ('-r', '--resend'):
            resend = True

    if config_file is None or xlsx_file is None:
        print example
    else:
        send_confirmation_email(config_file, xlsx_file, resend)
